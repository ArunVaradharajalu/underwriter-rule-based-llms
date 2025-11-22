"""
Test Harness Generator - Creates comprehensive Excel test harness file
with hierarchical rules, test cases, and automated execution tracking
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import json


class TestHarnessGenerator:
    """Generates Excel-based test harness with hierarchical rules and test cases"""

    def __init__(self):
        # Define color scheme (shared across all instances)
        self.colors = {
            'header': 'FFC000',      # Orange
            'pass': '92D050',        # Green
            'fail': 'FF0000',        # Red
            'pending': 'FFFF00',     # Yellow
            'info': 'D9E1F2',        # Light Blue
            'summary': '4472C4'      # Dark Blue
        }

    def generate_test_harness(self,
                             hierarchical_rules: List[Dict[str, Any]],
                             test_cases: List[Dict[str, Any]],
                             bank_id: str,
                             policy_type: str,
                             output_path: str,
                             test_execution_results: List[Dict[str, Any]] = None) -> str:
        """
        Generate comprehensive test harness Excel file

        Args:
            hierarchical_rules: List of hierarchical rules (tree structure)
            test_cases: List of test cases
            bank_id: Bank identifier
            policy_type: Policy type identifier
            output_path: Path to save Excel file
            test_execution_results: Optional list of test execution results to populate the Excel

        Returns:
            Path to generated file
        """
        print(f"📊 Generating test harness for {bank_id}/{policy_type}...")
        print(f"DEBUG [TestHarness]: Input - hierarchical_rules: {len(hierarchical_rules)} rules, test_cases: {len(test_cases)} cases")

        # Create a fresh workbook for each generation (prevents corruption from reuse)
        self.wb = openpyxl.Workbook()
        print(f"DEBUG [TestHarness]: Created new workbook, default sheets: {self.wb.sheetnames}")

        # Note: Removed calculation mode setting as we're not using formulas anymore

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
            print(f"DEBUG [TestHarness]: Removed default 'Sheet', remaining sheets: {self.wb.sheetnames}")

        # Flatten hierarchical rules for easier processing
        flattened_rules = self._flatten_rules(hierarchical_rules)
        print(f"DEBUG [TestHarness]: Flattened {len(hierarchical_rules)} hierarchical rules to {len(flattened_rules)} total rules")

        # Create sheets
        try:
            print(f"DEBUG [TestHarness]: Creating sheet 1/5 - Hierarchical Rules...")
            self._create_hierarchical_rules_sheet(hierarchical_rules, flattened_rules)
            print(f"DEBUG [TestHarness]: ✓ Sheet 1/5 created successfully")

            print(f"DEBUG [TestHarness]: Creating sheet 2/5 - Test Cases...")
            self._create_test_cases_sheet(test_cases)
            print(f"DEBUG [TestHarness]: ✓ Sheet 2/5 created successfully")

            print(f"DEBUG [TestHarness]: Creating sheet 3/5 - Test Execution...")
            self._create_execution_template_sheet(flattened_rules, test_cases, test_execution_results)
            print(f"DEBUG [TestHarness]: ✓ Sheet 3/5 created successfully")

            print(f"DEBUG [TestHarness]: Creating sheet 4/5 - Coverage Summary...")
            self._create_coverage_summary_sheet(flattened_rules, test_cases, test_execution_results)
            print(f"DEBUG [TestHarness]: ✓ Sheet 4/5 created successfully")

            print(f"DEBUG [TestHarness]: Creating sheet 5/5 - Instructions...")
            self._create_instructions_sheet(bank_id, policy_type)
            print(f"DEBUG [TestHarness]: ✓ Sheet 5/5 created successfully")

            print(f"DEBUG [TestHarness]: All sheets created. Final sheet list: {self.wb.sheetnames}")
        except Exception as e:
            print(f"ERROR [TestHarness]: Failed during sheet creation: {e}")
            import traceback
            traceback.print_exc()
            raise

        # Save workbook and ensure it's properly closed
        try:
            print(f"DEBUG [TestHarness]: Saving workbook to: {output_path}")
            print(f"DEBUG [TestHarness]: Workbook has {len(self.wb.sheetnames)} sheets: {self.wb.sheetnames}")

            # Important: Save without data_only to preserve formulas
            self.wb.save(output_path)
            print(f"DEBUG [TestHarness]: ✓ Workbook saved successfully")

            # Explicitly close the workbook to flush all buffers
            # Note: wb.close() may not exist in all openpyxl versions, so use try/except
            try:
                self.wb.close()
                print(f"DEBUG [TestHarness]: ✓ Workbook closed successfully")
            except AttributeError:
                # Older versions of openpyxl don't have close()
                print(f"DEBUG [TestHarness]: wb.close() not available (older openpyxl version)")
                pass

            # Verify file was created
            import os
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"DEBUG [TestHarness]: ✓ File verified: {output_path} ({file_size} bytes)")
            else:
                print(f"ERROR [TestHarness]: File not found after save: {output_path}")

            print(f"✅ Test harness saved to: {output_path}")
        except Exception as e:
            print(f"❌ ERROR [TestHarness]: Failed to save test harness: {e}")
            print(f"ERROR [TestHarness]: Exception type: {type(e).__name__}")
            import traceback
            traceback.print_exc()
            raise

        return output_path

    def _flatten_rules(self, rules: List[Dict[str, Any]], parent_id: str = None, level: int = 0) -> List[Dict[str, Any]]:
        """Flatten hierarchical rules tree to list with level information"""
        flattened = []

        for rule in rules:
            rule_copy = {
                'id': rule.get('id', ''),
                'name': rule.get('name', ''),
                'description': rule.get('description', ''),
                'expected': rule.get('expected', ''),
                'actual': rule.get('actual', ''),
                'confidence': rule.get('confidence', 0),
                'passed': rule.get('passed'),
                'page_number': rule.get('page_number'),
                'clause_reference': rule.get('clause_reference', ''),
                'parent_id': parent_id,
                'level': level
            }
            flattened.append(rule_copy)

            # Process dependencies (children)
            if 'dependencies' in rule and rule['dependencies']:
                children = self._flatten_rules(rule['dependencies'], rule['id'], level + 1)
                flattened.extend(children)

        return flattened

    def _create_hierarchical_rules_sheet(self, tree_rules: List[Dict[str, Any]], flattened_rules: List[Dict[str, Any]]):
        """Create Hierarchical Rules sheet with visual indentation"""
        ws = self.wb.create_sheet("Hierarchical Rules", 0)

        # Headers
        headers = ['Rule ID', 'Level', 'Rule Name', 'Description', 'Expected', 'Page', 'Clause', 'Confidence', 'Status']
        self._write_header_row(ws, headers)

        # Data rows
        row = 2
        for rule in flattened_rules:
            indent = "  " * rule['level']  # Indentation based on level

            ws.cell(row, 1, rule['id'])
            ws.cell(row, 2, rule['level'])
            ws.cell(row, 3, f"{indent}{rule['name']}")  # Indented name
            ws.cell(row, 4, rule['description'])
            ws.cell(row, 5, rule['expected'])
            ws.cell(row, 6, rule['page_number'] if rule['page_number'] else 'N/A')
            ws.cell(row, 7, rule['clause_reference'])
            ws.cell(row, 8, rule['confidence'])

            # Status indicator - based on rule's passed field
            passed = rule.get('passed')
            if passed is True:
                status = 'Pass'
                status_color = self.colors['pass']
            elif passed is False:
                status = 'Fail'
                status_color = self.colors['fail']
            else:
                # passed is None or not set - rule hasn't been evaluated yet
                status = 'Pending'
                status_color = self.colors['pending']
            
            status_cell = ws.cell(row, 9, status)
            status_cell.fill = PatternFill(start_color=status_color, fill_type='solid')

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_test_cases_sheet(self, test_cases: List[Dict[str, Any]]):
        """Create Test Cases sheet"""
        ws = self.wb.create_sheet("Test Cases")

        # Headers
        headers = ['Test ID', 'Test Name', 'Category', 'Priority', 'Expected Decision',
                   'Applicant Data', 'Policy Data', 'Expected Reasons']
        self._write_header_row(ws, headers)

        # Data rows
        row = 2
        for idx, tc in enumerate(test_cases, start=1):
            ws.cell(row, 1, f"TC{idx:03d}")
            ws.cell(row, 2, tc.get('test_case_name', ''))
            ws.cell(row, 3, tc.get('category', 'positive'))

            # Priority with color coding
            priority = tc.get('priority', 1)
            priority_cell = ws.cell(row, 4, priority)
            if priority == 1:
                priority_cell.fill = PatternFill(start_color='FF0000', fill_type='solid')
                priority_cell.font = Font(color='FFFFFF', bold=True)
            elif priority == 2:
                priority_cell.fill = PatternFill(start_color='FFC000', fill_type='solid')

            ws.cell(row, 5, tc.get('expected_decision', ''))
            ws.cell(row, 6, json.dumps(tc.get('applicant_data', {}), indent=2))
            ws.cell(row, 7, json.dumps(tc.get('policy_data', {}), indent=2))

            reasons = tc.get('expected_reasons', [])
            ws.cell(row, 8, '\n'.join(reasons) if isinstance(reasons, list) else str(reasons))

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 50

        ws.freeze_panes = 'A2'

    def _create_execution_template_sheet(self, flattened_rules: List[Dict[str, Any]], test_cases: List[Dict[str, Any]],
                                        test_execution_results: List[Dict[str, Any]] = None):
        """Create Test Execution Template with test results if available"""
        ws = self.wb.create_sheet("Test Execution")

        # Create a lookup dictionary for test results by test_case_id
        results_by_test_id = {}
        if test_execution_results:
            print(f"DEBUG [TestExecution]: Received {len(test_execution_results)} test results")
            for result in test_execution_results:
                test_case_id = result.get('test_case_id')
                if test_case_id:
                    results_by_test_id[test_case_id] = result
                    print(f"DEBUG [TestExecution]: Mapped test_case_id={test_case_id} to result")
            print(f"DEBUG [TestExecution]: Total mapped results: {len(results_by_test_id)}")
        else:
            print(f"DEBUG [TestExecution]: No test execution results provided")

        # Headers - One row per test case (not per rule)
        headers = ['Test ID', 'Test Case Name', 'Expected Decision', 'Actual Decision', 
                   'Expected Reasons', 'Actual Reasons', 'Result', 'Execution Time', 'Executed By', 'Notes']
        self._write_header_row(ws, headers)

        # Generate ONE row per test case (not per rule)
        row = 2
        for tc_idx, tc in enumerate(test_cases, start=1):
            test_id = f"TC{tc_idx:03d}"
            test_case_id = tc.get('id')
            test_case_name = tc.get('test_case_name', 'Unknown Test Case')
            
            # Get test execution result for this test case
            test_result = results_by_test_id.get(test_case_id) if test_case_id else None
            print(f"DEBUG [TestExecution]: Row {row}, Test {test_id}, test_case_id={test_case_id}, Found result: {test_result is not None}")

            # Test ID
            ws.cell(row, 1, test_id)
            
            # Test Case Name
            ws.cell(row, 2, test_case_name)
            
            # Expected Decision
            expected_decision = tc.get('expected_decision', '')
            ws.cell(row, 3, expected_decision)
            
            # Actual Decision - populate from test results if available
            if test_result:
                actual_decision = test_result.get('actual_decision', '')
                actual_cell = ws.cell(row, 4, actual_decision)
                actual_cell.fill = PatternFill(start_color='E2EFDA', fill_type='solid')  # Light green
            else:
                actual_cell = ws.cell(row, 4, '')
                actual_cell.fill = PatternFill(start_color='FFFF00', fill_type='solid')  # Yellow
            
            # Expected Reasons
            expected_reasons = tc.get('expected_reasons', [])
            ws.cell(row, 5, ', '.join(expected_reasons) if expected_reasons else 'None')
            
            # Actual Reasons
            if test_result:
                actual_reasons = test_result.get('actual_reasons', [])
                ws.cell(row, 6, ', '.join(actual_reasons) if actual_reasons else 'None')
            else:
                ws.cell(row, 6, '')
            
            # Result (Passed/Failed) - populate from test results if available
            if test_result:
                test_passed = test_result.get('test_passed')
                pass_status = 'PASS' if test_passed else 'FAIL'
                passed_cell = ws.cell(row, 7, pass_status)

                # Color code based on pass/fail
                if test_passed:
                    passed_cell.fill = PatternFill(start_color=self.colors['pass'], fill_type='solid')
                else:
                    passed_cell.fill = PatternFill(start_color=self.colors['fail'], fill_type='solid')
            else:
                passed_cell = ws.cell(row, 7, 'Pending')
                passed_cell.fill = PatternFill(start_color=self.colors['pending'], fill_type='solid')

            # Execution Time
            if test_result:
                exec_time = test_result.get('execution_time_ms', '')
                ws.cell(row, 8, f"{exec_time}ms" if exec_time else '')
            else:
                ws.cell(row, 8, '')

            # Executed By
            ws.cell(row, 9, 'system' if test_result else '')

            # Notes - Add fail reason if test failed
            if test_result and not test_result.get('test_passed'):
                fail_reason = test_result.get('fail_reason', '')
                if fail_reason:
                    ws.cell(row, 10, fail_reason)
                else:
                    ws.cell(row, 10, '')
            else:
                ws.cell(row, 12, '')

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 10  # Test ID
        ws.column_dimensions['B'].width = 40  # Test Case Name
        ws.column_dimensions['C'].width = 20  # Expected Decision
        ws.column_dimensions['D'].width = 20  # Actual Decision
        ws.column_dimensions['E'].width = 50  # Expected Reasons
        ws.column_dimensions['F'].width = 50  # Actual Reasons
        ws.column_dimensions['G'].width = 12  # Result
        ws.column_dimensions['H'].width = 15  # Execution Time
        ws.column_dimensions['I'].width = 15  # Executed By
        ws.column_dimensions['J'].width = 50  # Notes

        ws.freeze_panes = 'A2'

    def _create_coverage_summary_sheet(self, flattened_rules: List[Dict[str, Any]], test_cases: List[Dict[str, Any]],
                                      test_execution_results: List[Dict[str, Any]] = None):
        """Create Coverage Summary sheet with statistics and charts"""
        ws = self.wb.create_sheet("Coverage Summary")

        # Calculate actual counts from test results if available
        total_executed = 0
        passed_count = 0
        failed_count = 0
        if test_execution_results:
            for result in test_execution_results:
                total_executed += 1
                if result.get('test_passed'):
                    passed_count += 1
                else:
                    failed_count += 1

        # Calculate pending: total test cases minus executed test cases
        pending_count = len(test_cases) - total_executed if test_execution_results else len(test_cases)
        pass_rate = (passed_count / total_executed * 100) if total_executed > 0 else 0.0

        # Title (removed merge_cells to avoid Excel corruption)
        title_cell = ws.cell(1, 1, "Test Coverage Summary")
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.colors['summary'], fill_type='solid')
        # Note: Merged cells can cause Excel corruption, so avoiding ws.merge_cells()

        # Statistics
        row = 3
        stats = [
            ('Total Rules', len(flattened_rules)),
            ('Total Test Cases', len(test_cases)),
            ('Total Test Executions', len(test_cases)),  # One execution per test case
            ('', ''),
            ('Rules by Level', ''),
        ]

        # Count rules by level
        level_counts = {}
        for rule in flattened_rules:
            level = rule['level']
            level_counts[level] = level_counts.get(level, 0) + 1

        for level in sorted(level_counts.keys()):
            stats.append((f'  Level {level} Rules', level_counts[level]))

        stats.append(('', ''))
        stats.append(('Test Cases by Category', ''))

        # Count test cases by category
        category_counts = {}
        for tc in test_cases:
            category = tc.get('category', 'positive')
            category_counts[category] = category_counts.get(category, 0) + 1

        for category in sorted(category_counts.keys()):
            stats.append((f'  {category.capitalize()}', category_counts[category]))

        # Write statistics
        for label, value in stats:
            label_cell = ws.cell(row, 1, label)
            if label and not label.startswith('  '):
                label_cell.font = Font(bold=True)

            if value != '':
                ws.cell(row, 2, value)

            row += 1

        # Add execution status summary (will be calculated from Test Execution sheet)
        row += 2
        ws.cell(row, 1, "Execution Status").font = Font(bold=True, size=12)
        row += 1

        # Execution status with actual counts from test results
        ws.cell(row, 1, "Total Executed")
        ws.cell(row, 2, total_executed)
        row += 1

        ws.cell(row, 1, "Passed")
        pass_cell = ws.cell(row, 2, passed_count)
        pass_cell.fill = PatternFill(start_color=self.colors['pass'], fill_type='solid')
        row += 1

        ws.cell(row, 1, "Failed")
        fail_cell = ws.cell(row, 2, failed_count)
        fail_cell.fill = PatternFill(start_color=self.colors['fail'], fill_type='solid')
        row += 1

        ws.cell(row, 1, "Pending")
        pending_cell = ws.cell(row, 2, pending_count)
        pending_cell.fill = PatternFill(start_color=self.colors['pending'], fill_type='solid')
        row += 1

        # Calculate pass rate percentage from actual results
        ws.cell(row, 1, "Pass Rate %")
        ws.cell(row, 2, round(pass_rate, 2))

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_instructions_sheet(self, bank_id: str, policy_type: str):
        """Create Instructions sheet for test harness users"""
        ws = self.wb.create_sheet("Instructions")

        instructions = f"""
TEST HARNESS INSTRUCTIONS
========================

Bank: {bank_id}
Policy Type: {policy_type}

OVERVIEW
--------
This test harness provides a comprehensive framework for testing underwriting rules with automated
pass/fail validation and coverage tracking.

SHEETS DESCRIPTION
------------------

1. HIERARCHICAL RULES
   - Complete tree of all underwriting rules with visual hierarchy
   - Shows rule IDs, descriptions, expected values, page numbers, and clause references
   - Status column shows overall rule validation status

2. TEST CASES
   - All generated test cases with input data and expected outcomes
   - Categories: positive, negative, boundary, edge_case
   - Priority: 1 (high/critical), 2 (medium), 3 (low)

3. TEST EXECUTION (MAIN WORKING SHEET)
   - Fill in "Actual Result" column (E) during test execution
   - Manually update "Passed" column (F): Enter PASS if Expected = Actual, FAIL otherwise
   - Record execution date, tester name, and notes
   - Color coding: Yellow = Pending input needed

4. COVERAGE SUMMARY
   - Test coverage statistics and metrics
   - Manually update execution counts based on test results
   - Shows pass/fail rates and execution progress

HOW TO USE
----------

STEP 1: Review Rules
   - Open "Hierarchical Rules" sheet
   - Understand the rule structure and requirements
   - Note page numbers and clause references for policy lookup

STEP 2: Review Test Cases
   - Open "Test Cases" sheet
   - Understand test scenarios and expected outcomes
   - Focus on high-priority (1) test cases first

STEP 3: Execute Tests
   - Open "Test Execution" sheet
   - For each test case:
     a. Run the test with given input data
     b. Record the actual result in column E
     c. Compare Expected (D) vs Actual (E) and enter PASS or FAIL in column F
     d. Fill in execution date and tester name
     e. Add notes for any issues or observations

STEP 4: Monitor Coverage
   - Check "Coverage Summary" for progress
   - Aim for 100% execution coverage
   - Target > 95% pass rate

TIPS
----
- Start with Priority 1 test cases
- Test positive cases before negative cases
- Document all failures with detailed notes
- Update actual values exactly as shown in system output
- Manually verify Expected vs Actual and mark PASS/FAIL consistently

MANUAL VERIFICATION
-------------------
Compare the Expected (column D) vs Actual (column E) values:
- If they match exactly: Enter "PASS" in column F
- If they don't match: Enter "FAIL" in column F
- Update the Coverage Summary counts as you complete tests

For questions or issues, contact the testing team.
"""

        # Write instructions - explicitly set as text to prevent formula interpretation
        row = 1
        for line in instructions.split('\n'):
            cell = ws.cell(row, 1)
            # Set as explicit string to prevent Excel from treating special chars as formulas
            cell.value = str(line)
            cell.data_type = 's'  # 's' = string type, prevents formula interpretation
            row += 1

        # Format
        ws.column_dimensions['A'].width = 100
        for row_iter in ws.iter_rows(min_row=1, max_row=row):
            for cell in row_iter:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    def _write_header_row(self, ws, headers: List[str]):
        """Write and format header row"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def update_excel_with_test_results(self,
                                       excel_path: str,
                                       test_execution_results: List[Dict[str, Any]]) -> str:
        """
        Update existing Excel file with test execution results

        Args:
            excel_path: Path to existing Excel file
            test_execution_results: List of test execution results from TestExecutor

        Returns:
            Path to updated Excel file
        """
        print(f"Updating Excel file with test results: {excel_path}")

        # Load existing workbook
        wb = openpyxl.load_workbook(excel_path)

        if "Test Execution" not in wb.sheetnames:
            print("  ⚠ Warning: 'Test Execution' sheet not found in workbook")
            wb.close()
            return excel_path

        ws = wb["Test Execution"]

        # Build lookup by test_case_id (same as generate_test_harness method)
        results_by_case_id = {}
        for result in test_execution_results:
            test_case_id = result.get('test_case_id')
            if test_case_id:
                results_by_case_id[test_case_id] = result

        print(f"  Found {len(results_by_case_id)} test results to update")

        # Build test ID to result mapping by reading Test Cases sheet
        test_id_to_result = {}
        if "Test Cases" in wb.sheetnames:
            tc_ws = wb["Test Cases"]
            for tc_row in range(2, tc_ws.max_row + 1):
                test_id = tc_ws.cell(tc_row, 1).value  # Column A: Test ID (TC001, TC002, etc.)

                # Find matching result by test case name
                test_name = tc_ws.cell(tc_row, 2).value  # Column B: Test Name
                for result in test_execution_results:
                    if result.get('test_case_name') == test_name:
                        test_id_to_result[test_id] = result
                        break

        # Update each row in Test Execution sheet
        updated_count = 0
        for row in range(2, ws.max_row + 1):
            test_id = ws.cell(row, 1).value

            if test_id and test_id in test_id_to_result:
                result = test_id_to_result[test_id]

                # Column E: Actual Result
                actual_decision = result.get('actual_decision', '')
                actual_cell = ws.cell(row, 5, actual_decision)
                actual_cell.fill = PatternFill(start_color='E2EFDA', fill_type='solid')

                # Column F: Passed
                test_passed = result.get('test_passed', False)
                pass_status = 'PASS' if test_passed else 'FAIL'
                passed_cell = ws.cell(row, 6, pass_status)

                if test_passed:
                    passed_cell.fill = PatternFill(start_color=self.colors['pass'], fill_type='solid')
                else:
                    passed_cell.fill = PatternFill(start_color=self.colors['fail'], fill_type='solid')

                # Column G: Execution Date
                execution_time = result.get('execution_time_ms', '')
                ws.cell(row, 7, f"{execution_time}ms" if execution_time else '')

                # Column H: Executed By
                ws.cell(row, 8, 'system')

                # Column I: Notes (fail reason)
                fail_reason = result.get('fail_reason', '')
                if fail_reason and not test_passed:
                    ws.cell(row, 9, fail_reason)

                updated_count += 1

        # Save updated workbook
        wb.save(excel_path)
        wb.close()

        print(f"  ✓ Excel file updated: {updated_count} rows updated with test results")

        return excel_path
